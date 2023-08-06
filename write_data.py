import os
import pandas as pd
from openpyxl import load_workbook


def write_data(results):
    for result in results:
        if not isinstance(result, dict) or 'id' not in result or 'content' not in result:
            raise ValueError("Each message should be a dictionary with 'id' and 'content' keys")

    df_new = pd.DataFrame(results)
    file_path = 'result1.xlsx'

    if os.path.exists(file_path):
        # book = load_workbook(file_path)
        # writer = pd.ExcelWriter(file_path, engine='openpyxl')  # added mode='a' to append
        # writer.book = book
        # writer.sheets = {ws.title: ws for ws in book.worksheets}
        #
        # # writer = pd.ExcelWriter('test.xlsx', engine='openpyxl')
        # # writer.book = book
        # # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        writer = pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay")
        df_new.to_excel(writer, sheet_name="Sheet1")
    else:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')

    # writer = with pd.ExcelWriter("path_to_file.xlsx", mode="a", engine="openpyxl") as writer:
    # df.to_excel(writer, sheet_name="Sheet3")

    try:
        startrow = writer.sheets['Sheet1'].max_row
    except KeyError:
        writer.book.create_sheet('Sheet1')
        startrow = 0

    df_new.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=startrow)
    writer.close()
    print("=============================================")
    print("STATUS: DONE")
